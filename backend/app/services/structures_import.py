from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Iterator, Literal, Sequence
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from pydantic import EmailStr, TypeAdapter

from app.models.structure import (
    FirePolicy,
    StructureOpenPeriodKind,
    StructureOpenPeriodSeason,
    StructureType,
    WaterSource,
)
from app.models.availability import StructureUnit

EMAIL_ADAPTER = TypeAdapter(EmailStr)

HEADERS = [
    "name",
    "slug",
    "province",
    "address",
    "latitude",
    "longitude",
    "altitude",
    "type",
    "indoor_beds",
    "indoor_bathrooms",
    "indoor_showers",
    "indoor_activity_rooms",
    "has_kitchen",
    "hot_water",
    "land_area_m2",
    "shelter_on_field",
    "water_sources",
    "electricity_available",
    "fire_policy",
    "access_by_car",
    "access_by_coach",
    "access_by_public_transport",
    "coach_turning_area",
    "nearest_bus_stop",
    "weekend_only",
    "has_field_poles",
    "pit_latrine_allowed",
    "contact_emails",
    "website_urls",
    "notes_logistics",
    "notes",
]

OPEN_PERIOD_HEADERS = [
    "structure_slug",
    "kind",
    "season",
    "units",
    "date_start",
    "date_end",
    "notes",
]


TemplateFormat = Literal["xlsx", "csv"]


TEMPLATE_SAMPLE_ROWS: list[dict[str, object]] = [
    {
        "name": "Casa Alpina",
        "slug": "casa-alpina",
        "province": "TN",
        "address": "Via Bosco 10",
        "latitude": Decimal("46.123"),
        "longitude": Decimal("11.456"),
        "altitude": Decimal("1250"),
        "type": "house",
        "indoor_beds": 48,
        "indoor_bathrooms": 6,
        "indoor_showers": 10,
        "indoor_activity_rooms": 4,
        "has_kitchen": True,
        "hot_water": True,
        "land_area_m2": None,
        "shelter_on_field": False,
        "water_sources": None,
        "electricity_available": True,
        "fire_policy": "with_permit",
        "access_by_car": True,
        "access_by_coach": True,
        "access_by_public_transport": True,
        "coach_turning_area": True,
        "nearest_bus_stop": "Fermata Centro Scout",
        "weekend_only": False,
        "has_field_poles": False,
        "pit_latrine_allowed": False,
        "contact_emails": "info@example.org",
        "website_urls": "https://example.org/casa-alpina",
        "notes_logistics": "Accesso anche con pullman",
        "notes": "Spazi esterni ampi",
    },
    {
        "name": "Terreno Pianura",
        "slug": "terreno-pianura",
        "province": "MI",
        "address": "Località Campi",
        "latitude": Decimal("45.45"),
        "longitude": Decimal("9.12"),
        "altitude": Decimal("120"),
        "type": "land",
        "indoor_beds": None,
        "indoor_bathrooms": None,
        "indoor_showers": None,
        "indoor_activity_rooms": None,
        "has_kitchen": False,
        "hot_water": False,
        "land_area_m2": Decimal("5000"),
        "shelter_on_field": True,
        "water_sources": "tap",
        "electricity_available": False,
        "fire_policy": "allowed",
        "access_by_car": True,
        "access_by_coach": False,
        "access_by_public_transport": False,
        "coach_turning_area": False,
        "nearest_bus_stop": None,
        "weekend_only": True,
        "has_field_poles": True,
        "pit_latrine_allowed": True,
        "contact_emails": "prenotazioni@example.org;info@example.org",
        "website_urls": "https://example.org/terreno",
        "notes_logistics": "Campo estivo disponibile da giugno a agosto",
        "notes": "Ideale per campi estivi",
    },
]


OPEN_PERIOD_TEMPLATE_SAMPLE_ROWS: list[dict[str, object]] = [
    {
        "structure_slug": "casa-alpina",
        "kind": "season",
        "season": "summer",
        "units": "ALL",
        "date_start": None,
        "date_end": None,
        "notes": "Chiuso a ferragosto",
    },
    {
        "structure_slug": "terreno-pianura",
        "kind": "range",
        "season": None,
        "units": "EG,RS",
        "date_start": date(2025, 6, 1),
        "date_end": date(2025, 8, 31),
        "notes": "Campo estivo",
    },
]


@dataclass(slots=True)
class RowError:
    row: int
    field: str
    message: str
    source_format: TemplateFormat


@dataclass(slots=True)
class StructureImportRow:
    row: int
    name: str
    slug: str
    province: str
    address: str | None
    latitude: Decimal | None
    longitude: Decimal | None
    altitude: Decimal | None
    type: StructureType
    indoor_beds: int | None
    indoor_bathrooms: int | None
    indoor_showers: int | None
    indoor_activity_rooms: int | None
    has_kitchen: bool | None
    hot_water: bool | None
    land_area_m2: Decimal | None
    shelter_on_field: bool | None
    water_sources: list[WaterSource] | None
    electricity_available: bool | None
    fire_policy: FirePolicy | None
    access_by_car: bool | None
    access_by_coach: bool | None
    access_by_public_transport: bool | None
    coach_turning_area: bool | None
    nearest_bus_stop: str | None
    weekend_only: bool | None
    has_field_poles: bool | None
    pit_latrine_allowed: bool | None
    contact_emails: list[str]
    website_urls: list[str]
    notes_logistics: str | None
    notes: str | None


@dataclass(slots=True)
class ParsedWorkbook:
    rows: list[StructureImportRow]
    errors: list[RowError]
    blank_rows: int
    source_format: TemplateFormat


@dataclass(slots=True)
class StructureOpenPeriodImportRow:
    row: int
    structure_slug: str
    kind: StructureOpenPeriodKind
    season: StructureOpenPeriodSeason | None
    units: list[StructureUnit] | None
    date_start: date | None
    date_end: date | None
    notes: str | None


@dataclass(slots=True)
class ParsedOpenPeriods:
    rows: list[StructureOpenPeriodImportRow]
    errors: list[RowError]
    blank_rows: int
    source_format: TemplateFormat


def _normalise_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_units(value: object) -> tuple[list[StructureUnit] | None, list[str]]:
    text = _normalise_text(value)
    if not text:
        return None, []
    separators = re.compile(r"[;,]")
    candidates = [item.strip().upper() for item in separators.split(text) if item.strip()]
    units: list[StructureUnit] = []
    invalid: list[str] = []
    for candidate in candidates:
        try:
            units.append(StructureUnit(candidate))
        except ValueError:
            invalid.append(candidate)
    if invalid:
        return None, invalid
    if not units:
        return None, []
    return units, []


def _normalise_decimal(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        return Decimal(str(value).strip())
    except Exception as exc:  # pragma: no cover - defensive
        raise ValueError("invalid number") from exc


def _validate_province(value: str) -> str:
    if len(value) != 2 or not value.isalpha():
        raise ValueError("must be 2 letters")
    return value.upper()


def _validate_slug(value: str) -> str:
    if not value:
        raise ValueError("cannot be empty")
    return value


def _validate_type(value: str) -> StructureType:
    try:
        return StructureType(value.lower())
    except ValueError as exc:
        raise ValueError("must be one of house, land, mixed") from exc


def _validate_positive_int(value: object, *, allow_empty: bool = True) -> int | None:
    text = _normalise_text(value)
    if not text:
        return None if allow_empty else 0
    try:
        number = int(text)
    except ValueError as exc:
        raise ValueError("must be an integer") from exc
    if number < 0:
        raise ValueError("must be zero or greater")
    return number


def _validate_bool(value: object) -> bool | None:
    text = _normalise_text(value)
    if not text:
        return None
    lowered = text.lower()
    truthy = {"true", "1", "yes", "y", "si", "sì"}
    falsy = {"false", "0", "no", "n"}
    if lowered in truthy:
        return True
    if lowered in falsy:
        return False
    raise ValueError("must be true or false")


def _validate_decimal_non_negative(value: object) -> Decimal | None:
    decimal_value = _normalise_decimal(value)
    if decimal_value is None:
        return None
    if decimal_value < Decimal("0"):
        raise ValueError("must be zero or greater")
    return decimal_value


def _parse_water_sources(value: object) -> tuple[list[WaterSource] | None, list[str]]:
    text = _normalise_text(value)
    if not text:
        return None, []
    separators = re.compile(r"[;,]")
    candidates = [item.strip().lower() for item in separators.split(text) if item.strip()]
    sources: list[WaterSource] = []
    invalid: list[str] = []
    for candidate in candidates:
        try:
            sources.append(WaterSource(candidate))
        except ValueError:
            invalid.append(candidate)
    if invalid:
        return None, invalid
    if not sources:
        return None, []
    # Preserve order while removing duplicates
    seen: set[WaterSource] = set()
    unique_sources: list[WaterSource] = []
    for source in sources:
        if source not in seen:
            seen.add(source)
            unique_sources.append(source)
    return unique_sources, []


def _validate_fire_policy(value: object) -> FirePolicy | None:
    text = _normalise_text(value)
    if not text:
        return None
    try:
        return FirePolicy(text.lower())
    except ValueError as exc:
        allowed = ", ".join(item.value for item in FirePolicy)
        raise ValueError(f"must be one of {allowed}") from exc


def _validate_short_text(value: object, *, max_length: int) -> str | None:
    text = _normalise_text(value)
    if not text:
        return None
    if len(text) > max_length:
        raise ValueError(f"must be at most {max_length} characters")
    return text


def _validate_url(value: object) -> str | None:
    text = _normalise_text(value)
    if not text:
        return None
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("must be a valid http or https URL")
    return text


def _parse_website_urls(value: object) -> tuple[list[str], list[str]]:
    if value is None:
        return [], []

    if isinstance(value, (list, tuple)):
        raw_items = value
    else:
        text = _normalise_text(value)
        if not text:
            return [], []
        raw_items = re.split(r"[\n;]", text)

    urls: list[str] = []
    errors: list[str] = []
    seen: set[str] = set()

    for item in raw_items:
        candidate = _normalise_text(item)
        if not candidate:
            continue
        try:
            validated = _validate_url(candidate)
        except ValueError as exc:
            errors.append(f"{candidate}: {exc}")
            continue
        if validated is None:
            continue
        if validated not in seen:
            seen.add(validated)
            urls.append(validated)

    return urls, errors


def _parse_contact_emails(value: object) -> tuple[list[str], list[str]]:
    if value is None:
        return [], []

    if isinstance(value, (list, tuple)):
        raw_items = value
    else:
        text = _normalise_text(value)
        if not text:
            return [], []
        raw_items = re.split(r"[\n;,]", text)

    emails: list[str] = []
    errors: list[str] = []
    seen: set[str] = set()

    for item in raw_items:
        candidate = _normalise_text(item)
        if not candidate:
            continue
        lowered = candidate.lower()
        if lowered in seen:
            continue
        try:
            validated = EMAIL_ADAPTER.validate_python(candidate)
        except ValueError as exc:
            errors.append(f"{candidate}: {exc}")
            continue
        seen.add(lowered)
        emails.append(str(validated))

    return emails, errors


def _validate_latitude(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    if value < Decimal("-90") or value > Decimal("90"):
        raise ValueError("must be between -90 and 90")
    return value


def _validate_longitude(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    if value < Decimal("-180") or value > Decimal("180"):
        raise ValueError("must be between -180 and 180")
    return value


def _validate_altitude(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    if value < Decimal("-500") or value > Decimal("9000"):
        raise ValueError("must be between -500 and 9000")
    return value


def _is_blank_row(values: Sequence[object]) -> bool:
    for value in values:
        text = _normalise_text(value)
        if text:
            return False
    return True


def _process_rows(
    rows: Iterator[tuple[int, Sequence[object]]],
    *,
    source_format: TemplateFormat,
    max_rows: int,
) -> ParsedWorkbook:
    processed_rows = 0
    stored_rows: list[StructureImportRow] = []
    errors: list[RowError] = []
    seen_slugs: dict[str, int] = {}
    blank_rows = 0

    for index, raw_values in rows:
        values = list(raw_values)
        if len(values) < len(HEADERS):
            values.extend([None] * (len(HEADERS) - len(values)))
        else:
            values = values[: len(HEADERS)]

        if _is_blank_row(values):
            blank_rows += 1
            continue

        processed_rows += 1
        if processed_rows > max_rows:
            raise ValueError(f"Too many rows. Maximum allowed is {max_rows}")

        name_value = values[0]
        slug_value = values[1]
        province_value = values[2]
        address_value = values[3]
        latitude_raw = values[4]
        longitude_raw = values[5]
        altitude_raw = values[6]
        type_raw = values[7]
        indoor_beds_raw = values[8]
        indoor_bathrooms_raw = values[9]
        indoor_showers_raw = values[10]
        indoor_activity_rooms_raw = values[11]
        has_kitchen_raw = values[12]
        hot_water_raw = values[13]
        land_area_raw = values[14]
        shelter_on_field_raw = values[15]
        water_sources_raw = values[16]
        electricity_available_raw = values[17]
        fire_policy_raw = values[18]
        access_by_car_raw = values[19]
        access_by_coach_raw = values[20]
        access_by_public_transport_raw = values[21]
        coach_turning_area_raw = values[22]
        nearest_bus_stop_raw = values[23]
        weekend_only_raw = values[24]
        has_field_poles_raw = values[25]
        pit_latrine_allowed_raw = values[26]
        contact_emails_raw = values[27]
        website_urls_raw = values[28]
        notes_logistics_raw = values[29]
        notes_raw = values[30]

        row_errors: list[RowError] = []
        row_warnings: list[RowError] = []

        name = _normalise_text(name_value)
        if not name:
            row_errors.append(
                RowError(
                    row=index, field="name", message="cannot be empty", source_format=source_format
                )
            )

        try:
            slug = _validate_slug(_normalise_text(slug_value))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="slug", message=str(exc), source_format=source_format)
            )
            slug = ""
        try:
            province = _validate_province(_normalise_text(province_value))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="province", message=str(exc), source_format=source_format)
            )
            province = ""

        address = _normalise_text(address_value) or None

        try:
            latitude = _validate_latitude(_normalise_decimal(latitude_raw))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="latitude", message=str(exc), source_format=source_format)
            )
            latitude = None

        try:
            longitude = _validate_longitude(_normalise_decimal(longitude_raw))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="longitude", message=str(exc), source_format=source_format)
            )
            longitude = None

        try:
            altitude = _validate_altitude(_normalise_decimal(altitude_raw))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="altitude", message=str(exc), source_format=source_format)
            )
            altitude = None

        try:
            structure_type = _validate_type(_normalise_text(type_raw))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="type", message=str(exc), source_format=source_format)
            )
            structure_type = StructureType.HOUSE

        try:
            indoor_beds = _validate_positive_int(indoor_beds_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="indoor_beds", message=str(exc), source_format=source_format)
            )
            indoor_beds = None

        try:
            indoor_bathrooms = _validate_positive_int(indoor_bathrooms_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="indoor_bathrooms", message=str(exc), source_format=source_format)
            )
            indoor_bathrooms = None

        try:
            indoor_showers = _validate_positive_int(indoor_showers_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="indoor_showers", message=str(exc), source_format=source_format)
            )
            indoor_showers = None

        try:
            indoor_activity_rooms = _validate_positive_int(indoor_activity_rooms_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="indoor_activity_rooms",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            indoor_activity_rooms = None

        try:
            has_kitchen = _validate_bool(has_kitchen_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="has_kitchen", message=str(exc), source_format=source_format)
            )
            has_kitchen = None

        try:
            hot_water = _validate_bool(hot_water_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="hot_water", message=str(exc), source_format=source_format)
            )
            hot_water = None

        try:
            land_area_m2 = _validate_decimal_non_negative(land_area_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="land_area_m2", message=str(exc), source_format=source_format)
            )
            land_area_m2 = None

        try:
            shelter_on_field = _validate_bool(shelter_on_field_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="shelter_on_field", message=str(exc), source_format=source_format)
            )
            shelter_on_field = None

        water_sources_list, water_sources_invalid = _parse_water_sources(water_sources_raw)
        water_sources: list[WaterSource] | None
        if water_sources_invalid:
            row_errors.append(
                RowError(
                    row=index,
                    field="water_sources",
                    message=f"Invalid values: {', '.join(water_sources_invalid)}",
                    source_format=source_format,
                )
            )
            water_sources = None
        else:
            water_sources = water_sources_list
            if water_sources:
                exclusive_sources = {WaterSource.NONE, WaterSource.UNKNOWN}
                selected_exclusive = [
                    source for source in water_sources if source in exclusive_sources
                ]
                if selected_exclusive and len(water_sources) > 1:
                    joined = ", ".join(f"'{item.value}'" for item in selected_exclusive)
                    prefix = "Values" if len(selected_exclusive) > 1 else "Value"
                    row_errors.append(
                        RowError(
                            row=index,
                            field="water_sources",
                            message=f"{prefix} {joined} cannot be combined with other entries",
                            source_format=source_format,
                        )
                    )

        try:
            electricity_available = _validate_bool(electricity_available_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="electricity_available",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            electricity_available = None

        try:
            fire_policy = _validate_fire_policy(fire_policy_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="fire_policy", message=str(exc), source_format=source_format)
            )
            fire_policy = None

        try:
            access_by_car = _validate_bool(access_by_car_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="access_by_car", message=str(exc), source_format=source_format)
            )
            access_by_car = None

        try:
            access_by_coach = _validate_bool(access_by_coach_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="access_by_coach", message=str(exc), source_format=source_format)
            )
            access_by_coach = None

        try:
            access_by_public_transport = _validate_bool(access_by_public_transport_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="access_by_public_transport",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            access_by_public_transport = None

        try:
            coach_turning_area = _validate_bool(coach_turning_area_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="coach_turning_area",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            coach_turning_area = None

        try:
            nearest_bus_stop = _validate_short_text(nearest_bus_stop_raw, max_length=255)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="nearest_bus_stop",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            nearest_bus_stop = None

        try:
            weekend_only = _validate_bool(weekend_only_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="weekend_only", message=str(exc), source_format=source_format)
            )
            weekend_only = None

        try:
            has_field_poles = _validate_bool(has_field_poles_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="has_field_poles", message=str(exc), source_format=source_format)
            )
            has_field_poles = None

        try:
            pit_latrine_allowed = _validate_bool(pit_latrine_allowed_raw)
        except ValueError as exc:
            row_errors.append(
                RowError(
                    row=index,
                    field="pit_latrine_allowed",
                    message=str(exc),
                    source_format=source_format,
                )
            )
            pit_latrine_allowed = None

        contact_emails, email_errors = _parse_contact_emails(contact_emails_raw)
        for message in email_errors:
            row_errors.append(
                RowError(row=index, field="contact_emails", message=message, source_format=source_format)
            )

        website_urls, url_errors = _parse_website_urls(website_urls_raw)
        for message in url_errors:
            row_errors.append(
                RowError(row=index, field="website_urls", message=message, source_format=source_format)
            )

        notes_logistics = _normalise_text(notes_logistics_raw) or None
        notes = _normalise_text(notes_raw) or None

        def _warn(field: str, message: str) -> None:
            row_warnings.append(
                RowError(row=index, field=field, message=message, source_format=source_format)
            )

        if structure_type == StructureType.HOUSE:
            if land_area_m2 is not None:
                _warn("land_area_m2", "Ignored for type=house")
                land_area_m2 = None
            if shelter_on_field:
                _warn("shelter_on_field", "Ignored for type=house")
                shelter_on_field = None
            if water_sources:
                _warn("water_sources", "Ignored for type=house")
                water_sources = None
            if electricity_available:
                _warn("electricity_available", "Ignored for type=house")
                electricity_available = None
            if fire_policy is not None:
                _warn("fire_policy", "Ignored for type=house")
                fire_policy = None
            if has_field_poles:
                _warn("has_field_poles", "Ignored for type=house")
                has_field_poles = None
            if pit_latrine_allowed:
                _warn("pit_latrine_allowed", "Ignored for type=house")
                pit_latrine_allowed = None

        if structure_type == StructureType.LAND:
            if indoor_beds is not None:
                _warn("indoor_beds", "Ignored for type=land")
                indoor_beds = None
            if indoor_bathrooms is not None:
                _warn("indoor_bathrooms", "Ignored for type=land")
                indoor_bathrooms = None
            if indoor_showers is not None:
                _warn("indoor_showers", "Ignored for type=land")
                indoor_showers = None
            if indoor_activity_rooms is not None:
                _warn("indoor_activity_rooms", "Ignored for type=land")
                indoor_activity_rooms = None
            if has_kitchen:
                _warn("has_kitchen", "Ignored for type=land")
                has_kitchen = None
            if hot_water:
                _warn("hot_water", "Ignored for type=land")
                hot_water = None

        if slug and slug in seen_slugs:
            row_errors.append(
                RowError(
                    row=index,
                    field="slug",
                    message="duplicate slug in file",
                    source_format=source_format,
                )
            )
        elif slug:
            seen_slugs[slug] = index

        errors.extend(row_errors)
        if row_errors:
            continue

        stored_rows.append(
            StructureImportRow(
                row=index,
                name=name,
                slug=slug,
                province=province,
                address=address,
                latitude=latitude,
                longitude=longitude,
                altitude=altitude,
                type=structure_type,
                indoor_beds=indoor_beds,
                indoor_bathrooms=indoor_bathrooms,
                indoor_showers=indoor_showers,
                indoor_activity_rooms=indoor_activity_rooms,
                has_kitchen=has_kitchen,
                hot_water=hot_water,
                land_area_m2=land_area_m2,
                shelter_on_field=shelter_on_field,
                water_sources=water_sources,
                electricity_available=electricity_available,
                fire_policy=fire_policy,
                access_by_car=access_by_car,
                access_by_coach=access_by_coach,
                access_by_public_transport=access_by_public_transport,
                coach_turning_area=coach_turning_area,
                nearest_bus_stop=nearest_bus_stop,
                weekend_only=weekend_only,
                has_field_poles=has_field_poles,
                pit_latrine_allowed=pit_latrine_allowed,
                contact_emails=contact_emails,
                website_urls=website_urls,
                notes_logistics=notes_logistics,
                notes=notes,
            )
        )
        errors.extend(row_warnings)

    return ParsedWorkbook(
        rows=stored_rows,
        errors=errors,
        blank_rows=blank_rows,
        source_format=source_format,
    )


def _process_open_period_rows(
    rows: Iterator[tuple[int, Sequence[object]]],
    *,
    source_format: TemplateFormat,
    max_rows: int,
) -> ParsedOpenPeriods:
    processed_rows = 0
    stored_rows: list[StructureOpenPeriodImportRow] = []
    errors: list[RowError] = []
    blank_rows = 0

    for index, raw_values in rows:
        values = list(raw_values)
        if len(values) < len(OPEN_PERIOD_HEADERS):
            values.extend([None] * (len(OPEN_PERIOD_HEADERS) - len(values)))
        else:
            values = values[: len(OPEN_PERIOD_HEADERS)]

        if _is_blank_row(values):
            blank_rows += 1
            continue

        processed_rows += 1
        if processed_rows > max_rows:
            raise ValueError(f"Too many rows. Maximum allowed is {max_rows}")

        slug_raw, kind_raw, season_raw, units_raw, date_start_raw, date_end_raw, notes_raw = values

        row_errors: list[RowError] = []

        try:
            structure_slug = _validate_slug(_normalise_text(slug_raw))
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="structure_slug", message=str(exc), source_format=source_format)
            )
            structure_slug = ""

        kind_text = _normalise_text(kind_raw).lower()
        try:
            kind = StructureOpenPeriodKind(kind_text)
        except ValueError as exc:
            row_errors.append(
                RowError(row=index, field="kind", message="must be 'season' or 'range'", source_format=source_format)
            )
            kind = StructureOpenPeriodKind.SEASON

        season: StructureOpenPeriodSeason | None = None
        units: list[StructureUnit] | None = None
        date_start: date | None = None
        date_end: date | None = None

        season_text = _normalise_text(season_raw).lower()
        if season_text:
            try:
                season = StructureOpenPeriodSeason(season_text)
            except ValueError as exc:
                row_errors.append(
                    RowError(row=index, field="season", message=str(exc), source_format=source_format)
                )
                season = None

        units_parsed, unit_errors = _parse_units(units_raw)
        if unit_errors:
            row_errors.append(
                RowError(
                    row=index,
                    field="units",
                    message=f"Invalid units: {', '.join(unit_errors)}",
                    source_format=source_format,
                )
            )
        else:
            units = units_parsed

        date_start_text = _normalise_text(date_start_raw)
        if date_start_text:
            try:
                date_start = date.fromisoformat(date_start_text)
            except ValueError as exc:
                row_errors.append(
                    RowError(row=index, field="date_start", message="Invalid date", source_format=source_format)
                )
                date_start = None

        date_end_text = _normalise_text(date_end_raw)
        if date_end_text:
            try:
                date_end = date.fromisoformat(date_end_text)
            except ValueError as exc:
                row_errors.append(
                    RowError(row=index, field="date_end", message="Invalid date", source_format=source_format)
                )
                date_end = None

        if kind is StructureOpenPeriodKind.SEASON:
            if season is None:
                row_errors.append(
                    RowError(
                        row=index,
                        field="season",
                        message="Season is required for kind=season",
                        source_format=source_format,
                    )
                )
            if date_start is not None or date_end is not None:
                row_errors.append(
                    RowError(
                        row=index,
                        field="date_start",
                        message="Dates must be empty for kind=season",
                        source_format=source_format,
                    )
                )
                date_start = None
                date_end = None
        else:  # range
            if season is not None:
                row_errors.append(
                    RowError(
                        row=index,
                        field="season",
                        message="Season must be empty for kind=range",
                        source_format=source_format,
                    )
                )
                season = None
            if date_start is None or date_end is None:
                row_errors.append(
                    RowError(
                        row=index,
                        field="date_start",
                        message="Both date_start and date_end are required for kind=range",
                        source_format=source_format,
                    )
                )
            elif date_start > date_end:
                row_errors.append(
                    RowError(
                        row=index,
                        field="date_start",
                        message="date_start cannot be after date_end",
                        source_format=source_format,
                    )
                )

        notes = _normalise_text(notes_raw) or None

        errors.extend(row_errors)
        if row_errors:
            continue

        stored_rows.append(
            StructureOpenPeriodImportRow(
                row=index,
                structure_slug=structure_slug,
                kind=kind,
                season=season,
                units=units,
                date_start=date_start,
                date_end=date_end,
                notes=notes,
            )
        )

    return ParsedOpenPeriods(
        rows=stored_rows,
        errors=errors,
        blank_rows=blank_rows,
        source_format=source_format,
    )


def parse_structures_xlsx(data: bytes, *, max_rows: int = 2000) -> ParsedWorkbook:
    if not data:
        raise ValueError("The uploaded file is empty")

    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        try:
            header_row = next(sheet.iter_rows(max_row=1, values_only=True))
        except StopIteration as exc:
            raise ValueError("The uploaded file is empty") from exc

        header = [str(cell) if cell is not None else "" for cell in header_row]
        if [item.strip() for item in header] != HEADERS:
            raise ValueError("Invalid header. Please use the provided template")

        rows = _process_rows(
            ((index, tuple(row or tuple())) for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)),
            source_format="xlsx",
            max_rows=max_rows,
        )
        return rows
    finally:
        workbook.close()


def parse_structures_csv(data: bytes, *, max_rows: int = 2000) -> ParsedWorkbook:
    if not data:
        raise ValueError("The uploaded file is empty")

    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc

    buffer = StringIO(text)
    reader = csv.reader(buffer, delimiter=",")

    try:
        header = next(reader)
    except StopIteration as exc:
        raise ValueError("The uploaded file is empty") from exc

    if [item.strip() for item in header] != HEADERS:
        raise ValueError("Invalid header. Please use the provided template")

    return _process_rows(
        ((index, row) for index, row in enumerate(reader, start=2)),
        source_format="csv",
        max_rows=max_rows,
    )


def parse_structures_file(
    data: bytes,
    *,
    source_format: TemplateFormat,
    max_rows: int = 2000,
) -> ParsedWorkbook:
    if source_format == "xlsx":
        return parse_structures_xlsx(data, max_rows=max_rows)
    if source_format == "csv":
        return parse_structures_csv(data, max_rows=max_rows)
    raise ValueError("Unsupported file format. Only CSV and XLSX templates are supported.")


def parse_structure_open_periods_xlsx(
    data: bytes,
    *,
    max_rows: int = 2000,
) -> ParsedOpenPeriods:
    if not data:
        raise ValueError("The uploaded file is empty")

    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        try:
            header_row = next(sheet.iter_rows(max_row=1, values_only=True))
        except StopIteration as exc:
            raise ValueError("The uploaded file is empty") from exc

        header = [str(cell) if cell is not None else "" for cell in header_row]
        if [item.strip() for item in header] != OPEN_PERIOD_HEADERS:
            raise ValueError("Invalid header. Please use the provided template")

        rows = _process_open_period_rows(
            (
                (
                    index,
                    tuple(row or tuple()),
                )
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
            ),
            source_format="xlsx",
            max_rows=max_rows,
        )
        return rows
    finally:
        workbook.close()


def parse_structure_open_periods_csv(
    data: bytes,
    *,
    max_rows: int = 2000,
) -> ParsedOpenPeriods:
    if not data:
        raise ValueError("The uploaded file is empty")

    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc

    buffer = StringIO(text)
    reader = csv.reader(buffer, delimiter=",")

    try:
        header = next(reader)
    except StopIteration as exc:
        raise ValueError("The uploaded file is empty") from exc

    if [item.strip() for item in header] != OPEN_PERIOD_HEADERS:
        raise ValueError("Invalid header. Please use the provided template")

    return _process_open_period_rows(
        ((index, row) for index, row in enumerate(reader, start=2)),
        source_format="csv",
        max_rows=max_rows,
    )


def parse_structure_open_periods_file(
    data: bytes,
    *,
    source_format: TemplateFormat,
    max_rows: int = 2000,
) -> ParsedOpenPeriods:
    if source_format == "xlsx":
        return parse_structure_open_periods_xlsx(data, max_rows=max_rows)
    if source_format == "csv":
        return parse_structure_open_periods_csv(data, max_rows=max_rows)
    raise ValueError("Unsupported file format. Only CSV and XLSX templates are supported.")


def build_structures_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Structures"
    sheet.append(HEADERS)
    for row in TEMPLATE_SAMPLE_ROWS:
        sheet.append([row.get(header) for header in HEADERS])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_structures_template_csv() -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for row in TEMPLATE_SAMPLE_ROWS:
        writer.writerow([row.get(header) for header in HEADERS])
    return output.getvalue()


def build_structure_open_periods_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "structure_open_periods"
    sheet.append(OPEN_PERIOD_HEADERS)
    for row in OPEN_PERIOD_TEMPLATE_SAMPLE_ROWS:
        sheet.append([row.get(header) for header in OPEN_PERIOD_HEADERS])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_structure_open_periods_template_csv() -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(OPEN_PERIOD_HEADERS)
    for row in OPEN_PERIOD_TEMPLATE_SAMPLE_ROWS:
        writer.writerow([row.get(header) for header in OPEN_PERIOD_HEADERS])
    return output.getvalue()


__all__ = [
    "HEADERS",
    "OPEN_PERIOD_HEADERS",
    "TemplateFormat",
    "TEMPLATE_SAMPLE_ROWS",
    "OPEN_PERIOD_TEMPLATE_SAMPLE_ROWS",
    "ParsedWorkbook",
    "ParsedOpenPeriods",
    "RowError",
    "StructureImportRow",
    "StructureOpenPeriodImportRow",
    "parse_structures_file",
    "parse_structures_csv",
    "parse_structures_xlsx",
    "parse_structure_open_periods_file",
    "parse_structure_open_periods_csv",
    "parse_structure_open_periods_xlsx",
    "build_structures_template_workbook",
    "build_structures_template_csv",
    "build_structure_open_periods_template_workbook",
    "build_structure_open_periods_template_csv",
]
