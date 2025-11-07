from __future__ import annotations

from collections.abc import Sequence
import csv
import json
from io import BytesIO, StringIO
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.quote import Quote


def quote_to_xlsx(quote: Quote) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preventivo"

    sheet.append(["Preventivo", f"#{quote.id}"])
    sheet.append(["Evento", quote.event_id])
    structure_name = getattr(quote.structure, "name", None) if quote.structure else None
    sheet.append(["Struttura", structure_name or quote.structure_id])
    sheet.append(["Scenario", quote.scenario.value])
    sheet.append(["Valuta", quote.currency])
    sheet.append([])

    sheet.append(["Voce", "Quantità", "Importo unitario", "Totale"])
    for entry in quote.breakdown:
        sheet.append(
            [
                entry.get("description"),
                entry.get("quantity"),
                entry.get("unit_amount"),
                entry.get("total"),
            ]
        )

    sheet.append([])
    totals = quote.totals or {}
    sheet.append(["Subtotale", None, None, totals.get("subtotal", 0)])
    sheet.append(["Utenze", None, None, totals.get("utilities", 0)])
    sheet.append(["Tassa di soggiorno", None, None, totals.get("city_tax", 0)])
    sheet.append(["Totale", None, None, totals.get("total", 0)])
    sheet.append([
        "Caparra prenotazione",
        None,
        None,
        totals.get("booking_deposit", 0),
    ])
    sheet.append([
        "Deposito cauzionale",
        None,
        None,
        totals.get("damage_deposit", 0),
    ])
    sheet.append(["Caparre totali", None, None, totals.get("deposit", 0)])

    for column in range(1, 5):
        sheet.column_dimensions[get_column_letter(column)].width = 20

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _format_cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return value


def rows_to_csv_stream(
    rows: Sequence[dict[str, Any]],
    headers: Sequence[str],
) -> Iterator[bytes]:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(headers))
    writer.writeheader()
    yield buffer.getvalue().encode("utf-8")
    buffer.seek(0)
    buffer.truncate(0)
    for row in rows:
        filtered = {_header: _format_cell_value(row.get(_header)) for _header in headers}
        writer.writerow(filtered)
        yield buffer.getvalue().encode("utf-8")
        buffer.seek(0)
        buffer.truncate(0)


def rows_to_json_stream(rows: Sequence[dict[str, Any]]) -> Iterator[bytes]:
    yield b"["
    for index, row in enumerate(rows):
        payload = json.dumps(row, ensure_ascii=False)
        if index:
            yield b"," + payload.encode("utf-8")
        else:
            yield payload.encode("utf-8")
    yield b"]"


def rows_to_xlsx_stream(
    rows: Sequence[dict[str, Any]],
    headers: Sequence[str],
) -> Iterator[bytes]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Export"

    sheet.append(list(headers))
    for row in rows:
        sheet.append([_format_cell_value(row.get(header)) for header in headers])

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 20

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    chunk = buffer.read(8192)
    while chunk:
        yield chunk
        chunk = buffer.read(8192)


__all__ = [
    "quote_to_xlsx",
    "rows_to_csv_stream",
    "rows_to_json_stream",
    "rows_to_xlsx_stream",
]
