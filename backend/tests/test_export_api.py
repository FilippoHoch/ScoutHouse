import csv
import json
import os
from collections.abc import Generator
from io import BytesIO
from openpyxl import load_workbook
from urllib.parse import quote
from uuid import uuid4
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///./test.db")
os.environ.setdefault("APP_ENV", "test")

from app.api.v1.export import (  # noqa: E402
    CSV_HEADERS_EVENTS,
    CSV_HEADERS_OPEN_PERIODS,
    CSV_HEADERS_STRUCTURES,
)
from app.core.db import Base, engine  # noqa: E402
from app.core.limiter import TEST_RATE_LIMIT_HEADER  # noqa: E402
from app.main import app  # noqa: E402
from tests.utils import auth_headers, create_user, participants_payload  # noqa: E402


@pytest.fixture(autouse=True)
def setup_database() -> Generator[None, None, None]:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def get_client(
    *, authenticated: bool = False, is_admin: bool = False, email: str | None = None
) -> TestClient:
    client = TestClient(app)
    if authenticated:
        if email is not None and email != "test@example.com":
            create_user(email=email, name="Alt User", is_admin=is_admin)
            response = client.post(
                "/api/v1/auth/login",
                json={"email": email, "password": "password123"},
                headers={TEST_RATE_LIMIT_HEADER: str(uuid4())},
            )
            assert response.status_code == 200
            token = response.json()["access_token"]
            client.headers.update({"Authorization": f"Bearer {token}"})
        else:
            client.headers.update(auth_headers(client, is_admin=is_admin))
    return client


def create_structure(
    client: TestClient,
    slug: str,
    province: str,
    *,
    open_periods: list[dict[str, object]] | None = None,
) -> None:
    payload = {
        "name": f"Structure {slug}",
        "slug": slug,
        "province": province,
        "type": "house",
        "address": "Via Rover 1",
        "latitude": 45.0,
        "longitude": 9.0,
        "altitude": 210.0,
    }
    if open_periods is not None:
        payload["open_periods"] = open_periods
    response = client.post("/api/v1/structures/", json=payload)
    assert response.status_code == 201


def create_event(
    client: TestClient,
    title: str,
    start: str,
    end: str,
    *,
    branch: str = "LC",
    status: str = "planning",
    budget_total: str | None = None,
) -> int:
    payload = {
        "title": title,
        "branch": branch,
        "start_date": start,
        "end_date": end,
        "participants": participants_payload(lc=10, leaders=2),
        "status": status,
    }
    if budget_total is not None:
        payload["budget_total"] = budget_total
    response = client.post("/api/v1/events/", json=payload)
    assert response.status_code == 201
    return response.json()["id"]


def test_export_structures_requires_admin() -> None:
    client = get_client(authenticated=True)
    response = client.get("/api/v1/export/structures?format=csv")
    assert response.status_code == 403


def test_export_structures_with_filters_csv() -> None:
    client = get_client(authenticated=True, is_admin=True)
    open_period = {
        "kind": "range",
        "date_start": "2025-06-01",
        "date_end": "2025-06-30",
        "notes": "Disponibile in estate",
        "units": ["LC", "EG"],
    }
    create_structure(client, "casa-scout", "MI", open_periods=[open_period])
    create_structure(client, "casa-nord", "BS")

    filters = quote(json.dumps({"province": "MI"}))
    response = client.get(f"/api/v1/export/structures?format=csv&filters={filters}")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="structures.zip"'

    with ZipFile(BytesIO(response.content)) as archive:
        assert set(archive.namelist()) == {"structures.csv", "structure_open_periods.csv"}
        structures = archive.read("structures.csv").decode("utf-8").strip().splitlines()
        assert len(structures) == 2  # header + single row
        header = next(csv.reader([structures[0]]))
        assert header[8] == "altitude"
        assert "casa-scout" in structures[1]
        assert all("casa-nord" not in line for line in structures)

        open_period_lines = archive.read("structure_open_periods.csv").decode("utf-8").splitlines()
        open_period_reader = list(csv.reader(open_period_lines))
        assert open_period_reader[0] == list(CSV_HEADERS_OPEN_PERIODS)
        data_rows = [row for row in open_period_reader[1:] if any(cell for cell in row)]
        empty_rows = [row for row in open_period_reader[1:] if not any(cell for cell in row)]
        for row in empty_rows:
            assert all(cell == "" for cell in row)
        assert data_rows, "Expected at least one exported open period"
        assert all(row[1] == "casa-scout" for row in data_rows)
        first_period_row = data_rows[0]
        assert first_period_row[2] == "range"
        assert first_period_row[4] == "LC,EG"
        assert first_period_row[5] == "2025-06-01"
        assert first_period_row[6] == "2025-06-30"
        assert first_period_row[7] == "Disponibile in estate"


def test_export_structures_xlsx_includes_open_periods_sheet() -> None:
    client = get_client(authenticated=True, is_admin=True)
    open_period = {
        "kind": "season",
        "season": "summer",
        "notes": "Aperto in estate",
        "units": ["ALL"],
    }
    create_structure(client, "casa-scout", "MI", open_periods=[open_period])

    response = client.get("/api/v1/export/structures?format=xlsx")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="structures.xlsx"'

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        assert set(workbook.sheetnames) == {"structures", "structure_open_periods"}
        structures_sheet = workbook["structures"]
        structure_rows = list(structures_sheet.iter_rows(values_only=True))
        assert tuple(structure_rows[0]) == CSV_HEADERS_STRUCTURES
        data_rows = [row for row in structure_rows[1:] if any(row)]
        assert data_rows, "Expected at least one exported structure"
        assert data_rows[0][1] == "casa-scout"

        periods_sheet = workbook["structure_open_periods"]
        period_rows = list(periods_sheet.iter_rows(values_only=True))
        assert tuple(period_rows[0]) == CSV_HEADERS_OPEN_PERIODS
        period_data = [row for row in period_rows[1:] if any(row)]
        assert period_data, "Expected open periods sheet to include data"
        assert period_data[0][1] == "casa-scout"
        assert period_data[0][2] == "season"
    finally:
        workbook.close()
def test_export_events_json_with_date_range() -> None:
    client = get_client(authenticated=True)
    first_event = create_event(client, "Campo Invernale", "2025-01-10", "2025-01-12")
    create_event(client, "Campo Estivo", "2025-07-01", "2025-07-05")

    response = client.get("/api/v1/export/events?format=json&from=2025-01-01&to=2025-01-31")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="events.json"'

    payload = json.loads(response.content.decode("utf-8"))
    assert isinstance(payload, list)
    assert len(payload) == 1
    assert payload[0]["id"] == first_event


def test_export_events_with_additional_filters() -> None:
    client = get_client(authenticated=True)
    matching = create_event(
        client,
        "Campo Invernale",
        "2025-02-10",
        "2025-02-15",
        branch="EG",
        status="booked",
        budget_total="1200.00",
    )
    create_event(
        client,
        "Uscita di Branca LC",
        "2025-02-12",
        "2025-02-14",
        branch="LC",
        status="planning",
    )
    create_event(
        client,
        "Campo Archiviato",
        "2024-08-01",
        "2024-08-05",
        branch="EG",
        status="archived",
        budget_total="900.00",
    )

    response = client.get(
        "/api/v1/export/events?format=json"
        "&from=2025-02-01&to=2025-02-28"
        "&q=Invernale&branch=EG&status=booked&budget=with"
    )
    assert response.status_code == 200

    events = json.loads(response.content.decode("utf-8"))
    assert isinstance(events, list)
    assert [event["id"] for event in events] == [matching]


def test_export_events_xlsx_contains_headers() -> None:
    client = get_client(authenticated=True)
    event_id = create_event(client, "Campo Invernale", "2025-02-10", "2025-02-12")

    response = client.get("/api/v1/export/events?format=xlsx")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="events.xlsx"'

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(max_row=1, values_only=True))
        assert tuple(header_row) == CSV_HEADERS_EVENTS
        data_rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)]
        assert data_rows, "Expected at least one exported event"
        assert data_rows[0][0] == event_id
    finally:
        workbook.close()


def test_event_ical_download_and_permissions() -> None:
    owner_client = get_client(authenticated=True)
    event_id = create_event(owner_client, "Riunione Team", "2025-03-01", "2025-03-02")

    response = owner_client.get(f"/api/v1/events/{event_id}/ical")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/calendar")
    body = response.text
    assert "BEGIN:VEVENT" in body
    assert "SUMMARY:Riunione Team" in body
    assert "DTSTART;VALUE=DATE:20250301" in body

    other_client = get_client(authenticated=True, email="alt@example.com")
    forbidden = other_client.get(f"/api/v1/events/{event_id}/ical")
    assert forbidden.status_code == 403
