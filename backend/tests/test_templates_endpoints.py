from __future__ import annotations

import csv
import json
import os
from collections.abc import Generator
from io import BytesIO, StringIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///./test.db")
os.environ.setdefault("APP_ENV", "test")

from app.core.db import Base, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.services.structures_import import HEADERS, OPEN_PERIOD_HEADERS  # noqa: E402


@pytest.fixture(autouse=True)
def setup_database() -> Generator[None, None, None]:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def get_client() -> TestClient:
    return TestClient(app)


def test_structures_template_xlsx_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structures.xlsx")
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structures_import_template.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(max_row=1, values_only=True))
    assert [str(cell) for cell in header_row] == HEADERS
    workbook.close()


def test_structures_template_csv_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structures.csv")
    assert response.status_code == 200
    assert response.headers["content-type"] == "text/csv; charset=utf-8"
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structures_import_template.csv"'
    )

    buffer = StringIO(response.content.decode("utf-8"))
    reader = csv.reader(buffer)
    header = next(reader)
    assert header == HEADERS


def test_structures_template_json_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structures.json")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/json; charset=utf-8"
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structures_import_template.json"'
    )

    payload = json.loads(response.content.decode("utf-8"))
    assert isinstance(payload, list)
    assert payload, "Expected sample rows"
    first_row = payload[0]
    assert isinstance(first_row, dict)
    for header in HEADERS:
        assert header in first_row


def test_structure_open_periods_template_xlsx_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structure-open-periods.xlsx")
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structure_open_periods_template.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(max_row=1, values_only=True))
    assert [str(cell) for cell in header_row] == OPEN_PERIOD_HEADERS
    workbook.close()


def test_structure_open_periods_template_csv_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structure-open-periods.csv")
    assert response.status_code == 200
    assert response.headers["content-type"] == "text/csv; charset=utf-8"
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structure_open_periods_template.csv"'
    )

    buffer = StringIO(response.content.decode("utf-8"))
    reader = csv.reader(buffer)
    header = next(reader)
    assert header == OPEN_PERIOD_HEADERS


def test_structure_open_periods_template_json_download() -> None:
    client = get_client()
    response = client.get("/api/v1/templates/structure-open-periods.json")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/json; charset=utf-8"
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="structure_open_periods_template.json"'
    )

    payload = json.loads(response.content.decode("utf-8"))
    assert isinstance(payload, list)
    assert payload, "Expected sample rows"
    first_row = payload[0]
    assert isinstance(first_row, dict)
    for header in OPEN_PERIOD_HEADERS:
        assert header in first_row
